import json
import pandas as pd
from io import BytesIO
from decimal import Decimal
from openpyxl import Workbook
from datetime import datetime
from django.conf import settings
from .forms import MovimientoForm
from django.contrib import messages
from django.http import HttpResponse
from django.core.mail import send_mail
from django.contrib.auth.models import User
from django.forms.models import model_to_dict
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth import authenticate, login, logout
from .models import IngresoMensual, MovimientoLog, SystemLog
from django.shortcuts import render, redirect, get_object_or_404
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.contrib.auth.decorators import login_required, user_passes_test
from .forms import CustomUserCreationForm, CustomLoginForm, ProfileUpdateForm

# --- CONSULTAR REGISTROS Y ORDENAR CORRECTAMENTE ---
registros = list(IngresoMensual.objects.all().values())

# --- LOGIN ---
def login_view(request):
    if request.user.is_authenticated:
        return redirect('index')

    form = CustomLoginForm(request, data=request.POST or None)
    if request.method == "POST" and form.is_valid():
        user = form.get_user()
        login(request, user)
        return redirect('index')

    return render(request, "finanzas/login.html", {"form": form})

# --- LOGOUT ---
@login_required
def logout_view(request):
    logout(request)
    return redirect('login')

def index(request):
    form = MovimientoForm(request.POST or None)
    mensaje = ""

    # --- AÑADIR MOVIMIENTO ---
    if request.method == "POST" and "añadir" in request.POST and form.is_valid():
        periodo_existente = form.cleaned_data["periodo"].periodo if form.cleaned_data["periodo"] else None
        nuevo_periodo = form.cleaned_data["nuevo_periodo"]
        columna = form.cleaned_data["columna"]
        monto = Decimal(form.cleaned_data["monto"])

        periodo_final = nuevo_periodo.strip() if nuevo_periodo else periodo_existente
        ingreso, _ = IngresoMensual.objects.get_or_create(periodo=periodo_final)
        valor_actual = getattr(ingreso, columna, 0) or 0
        setattr(ingreso, columna, valor_actual + monto)
        ingreso.save()

        # --- Registrar log ---
        MovimientoLog.objects.create(
            usuario=request.user,
            tipo="añadir",
            periodo=periodo_final,
            columna=columna,
            monto=monto,
            observaciones=f"Añadido {monto} a '{columna}' en {periodo_final}"
        )

        mensaje = f"✅ Se añadió {monto:,.2f} a '{columna}' en {periodo_final}"

    # --- ELIMINAR REGISTRO ---
    if request.method == "POST" and "eliminar" in request.POST:
        id_registro = request.POST.get("id_registro")
        ingreso = get_object_or_404(IngresoMensual, id=id_registro)
        MovimientoLog.objects.create(
            usuario=request.user,
            tipo="eliminar",
            periodo=ingreso.periodo,
            columna="N/A",
            monto=0,
            observaciones=f"Eliminado registro ID {id_registro} ({ingreso.periodo})"
        )
        ingreso.delete()
        mensaje = f"Registro {id_registro} eliminado correctamente."

    # --- EDITAR REGISTRO ---
    if request.method == "POST" and "editar" in request.POST:
        id_registro = request.POST.get("id_registro")
        columna = request.POST.get("columna")
        nuevo_valor = Decimal(request.POST.get("nuevo_valor", 0))
        ingreso = get_object_or_404(IngresoMensual, id=id_registro)
        setattr(ingreso, columna, nuevo_valor)
        ingreso.save()

        # --- Registrar log ---
        MovimientoLog.objects.create(
            usuario=request.user,
            tipo="editar",
            periodo=ingreso.periodo,
            columna=columna,
            monto=nuevo_valor,
            observaciones=f"Editado '{columna}' en {ingreso.periodo}"
        )

        mensaje = f"Registro {id_registro} actualizado correctamente."

    # --- GENERAR REPORTE ---
    if request.method == "POST" and "generar_reporte" in request.POST:
        inicio = request.POST.get("inicio")
        fin = request.POST.get("fin")
        try:
            inicio_obj = IngresoMensual.objects.get(periodo=inicio)
            fin_obj = IngresoMensual.objects.get(periodo=fin)
        except IngresoMensual.DoesNotExist:
            return HttpResponse("Alguno de los periodos no existe.")

        data = IngresoMensual.objects.filter(
            id__gte=inicio_obj.id,
            id__lte=fin_obj.id
        ).order_by("id")
        if not data.exists():
            return HttpResponse("No hay datos para ese rango.")

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Financiero"
        headers = [
            "Periodo", "Ingresos x  Mantenimiento", "DPPP", "Ingresos Netos por Mantenimiento", "Ingreso x Cuota Extraordinaria",
            "Cuota ordinaria retroactiva", "Revision CSAU", "Depósitos en garantia de obra", "Ingresos x Intereses de Cuotas",
            "Ingresos por Rendimiento de Inversiones", "Sanciones", "Recuperación de Seguro/daños", "Recuperación de gastos por Cobranza via legal", "Depositos no identificados",
            "Total", "Ingresos reales vs fact", "Diferencia de ingresos fac vs cobrados", "Observaciones"
        ]
        ws.append(headers)
        header_fill = PatternFill("solid", fgColor="1e3a8a")
        header_font = Font(color="FFFFFF", bold=True)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for r in data:
            ws.append([
                r.periodo, r.ingresos_mantenimiento, r.dppp, r.ingresos_netos_mantenimiento,
                r.ingresos_cuota_extraordinaria, r.cuota_ordinaria_retroactiva,
                r.revision_csau, r.depositos_garantia_obra,
                r.ingresos_intereses_cuotas, r.ingresos_rendimiento_inversiones,
                r.sanciones, r.recuperacion_seguro_danios,
                r.recuperacion_gastos_cobranza, r.depositos_no_identificados,
                r.total(), r.ingresos_reales_vs_fact, r.diferencia_ingresos_fac_vs_cobrados
            ])

        # Fila de totales
        total_row = ["TOTAL"]
        for col in range(2, len(headers) + 1):
            col_letter = ws.cell(row=1, column=col).column_letter
            total_row.append(f"=SUM({col_letter}2:{col_letter}{len(data) + 1})")
        ws.append(total_row)
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="eab308")

        border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC")
        )
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename=\"reporte_finanzas.xlsx\"'
        return response

    # --- CONSULTA FINAL ---
    registros = IngresoMensual.objects.all().order_by("id")
    periodos = IngresoMensual.objects.values_list("periodo", flat=True).distinct().order_by("id")

    # --- FILTRO POR PERIODO ---
    inicio = request.GET.get("inicio")
    fin = request.GET.get("fin")
    registros_qs = IngresoMensual.objects.all().order_by("id")

    if inicio and fin:
        try:
            inicio_obj = IngresoMensual.objects.get(periodo=inicio)
            fin_obj = IngresoMensual.objects.get(periodo=fin)
            registros_qs = registros_qs.filter(id__gte=inicio_obj.id, id__lte=fin_obj.id)
        except IngresoMensual.DoesNotExist:
            registros_qs = IngresoMensual.objects.none()  # No existe alguno de los periodos
    elif inicio:  # solo inicio
        try:
            inicio_obj = IngresoMensual.objects.get(periodo=inicio)
            registros_qs = registros_qs.filter(id__gte=inicio_obj.id)
        except IngresoMensual.DoesNotExist:
            registros_qs = IngresoMensual.objects.none()
    elif fin:  # solo fin
        try:
            fin_obj = IngresoMensual.objects.get(periodo=fin)
            registros_qs = registros_qs.filter(id__lte=fin_obj.id)
        except IngresoMensual.DoesNotExist:
            registros_qs = IngresoMensual.objects.none()

    registros = list(registros_qs)

    # --- CÁLCULOS PARA TARJETAS KPI ---
    total_mantenimiento = sum((r.ingresos_mantenimiento or Decimal(0)) for r in registros)
    total_dppp = sum((r.dppp or Decimal(0)) for r in registros)
    total_netos = sum((r.ingresos_netos_mantenimiento or Decimal(0)) for r in registros)
    promedio_diferencia = (
        sum((r.diferencia_ingresos_fac_vs_cobrados or Decimal(0)) for r in registros) / len(registros)
        if registros else 0
)
    
    # Convierte registros en lista de diccionarios para el JS
    registros_serializados = [model_to_dict(r) for r in registros]
    registros_json = json.dumps(registros_serializados, default=str)

    return render(request, "finanzas/index.html", {
        "form": form,
        "mensaje": mensaje,
        "registros": registros,
        "registros_json": registros_json,
        "periodos": periodos,
        "total_mantenimiento": total_mantenimiento,
        "total_dppp": total_dppp,
        "total_netos": total_netos,
        "promedio_diferencia": promedio_diferencia,
    })

@login_required
def historial_movimientos(request):
    movimientos = MovimientoLog.objects.select_related("usuario").all().order_by("-fecha")

    if request.method == "POST" and "eliminar_mov" in request.POST:
        mov_id = request.POST.get("id_mov")
        movimiento = get_object_or_404(MovimientoLog, id=mov_id)

        # --- Ajustar el IngresoMensual solo si es tipo "añadir" ---
        if movimiento.tipo == "añadir" and movimiento.columna and movimiento.monto:
            try:
                ingreso = IngresoMensual.objects.get(periodo=movimiento.periodo)
                valor_actual = getattr(ingreso, movimiento.columna, 0) or 0
                setattr(ingreso, movimiento.columna, max(valor_actual - movimiento.monto, 0))
                
                # Recalcular diferencia automáticamente
                ingreso.ingresos_netos_mantenimiento = (
                    (ingreso.ingresos_mantenimiento or 0) - (ingreso.dppp or 0)
                )
                ingreso.save()
            except IngresoMensual.DoesNotExist:
                pass  # No hay ingreso para ese periodo, no hacemos nada

        # --- Registrar log en SystemLog ---
        SystemLog.objects.create(
            usuario=request.user,
            accion="eliminar_movimiento",
            detalle=f"El usuario {request.user.username} eliminó movimiento ID {mov_id} (Tipo: {movimiento.tipo}, Periodo: {movimiento.periodo})"
        )

        # --- Eliminar movimiento ---
        movimiento.delete()
        messages.success(request, "Movimiento eliminado correctamente.")
        return redirect("historial_movimientos")

    return render(request, "finanzas/historial_movimientos.html", {
        "movimientos": movimientos
    })


@login_required
def profile(request):
    user = request.user
    password_form = PasswordChangeForm(user=user)
    profile_form = ProfileUpdateForm(instance=user)

    if request.method == "POST":
        if "update_profile" in request.POST:
            profile_form = ProfileUpdateForm(request.POST, instance=user)
            if profile_form.is_valid():
                profile_form.save()
                registrar_log(user, "Actualización de perfil", f"Correo actualizado a {profile_form.cleaned_data.get('email')}")
                messages.success(request, "Perfil actualizado correctamente.")
                return redirect("profile")

        elif "change_password" in request.POST:
            password_form = PasswordChangeForm(user=user, data=request.POST)
            if password_form.is_valid():
                password_form.save()
                update_session_auth_hash(request, password_form.user)  # evita que cierre sesión
                registrar_log(user, "Cambio de contraseña", "El usuario cambió su contraseña.")
                messages.success(request, "Contraseña actualizada correctamente.")
                return redirect("profile")

    return render(
        request,
        "finanzas/profile.html",
        {
            "profile_form": profile_form,
            "password_form": password_form,
        },
    )

# --- ADMIN USERS ---
@login_required
@user_passes_test(lambda u: u.is_staff)
def admin_users(request):
    form = CustomUserCreationForm()

    if request.method == "POST":
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            # Crear usuario
            user = form.save(commit=False)
            user.save()

            # Registrar log de creación
            registrar_log(request.user, "Creación de usuario", f"Se creó el usuario '{user.username}'.")

            # Intentar enviar correo
            try:
                send_mail(
                    subject="Datos de acceso - Panel Financiero",
                    message=(
                        f"Hola {user.username},\n\n"
                        f"Tu cuenta ha sido creada exitosamente.\n\n"
                        f"Usuario: {user.username}\n"
                        f"Correo: {user.email}\n"
                        f"Contraseña: (la que elegiste)\n\n"
                        f"Por favor, cambia tu contraseña al ingresar."
                    ),
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[user.email],
                    fail_silently=False,
                )
                messages.success(request, f"Usuario '{user.username}' creado y correo enviado correctamente.")
            except Exception as e:
                messages.warning(request, f"Usuario '{user.username}' creado, pero no se pudo enviar el correo. ({e})")

            # Limpiar formulario
            form = CustomUserCreationForm()
            return redirect("admin_users")

        else:
            # Cuando el formulario NO es válido, no existe 'user'
            messages.error(request, "Error al crear el usuario. Revisa los campos.")
            correo = form.cleaned_data.get("email", "sin email")
            registrar_log(request.user, "Error al crear usuario", f"No se pudo crear el usuario con correo '{correo}'")

    users = User.objects.all().order_by('-id')
    return render(request, "finanzas/admin_users.html", {"form": form, "users": users})

# --- ADMIN CONFIG ---
@login_required
@user_passes_test(lambda u: u.is_staff)
def admin_config(request):
    message = ""

    if request.method == "POST":
        nombre_sistema = request.POST.get("nombre_sistema")
        moneda = request.POST.get("moneda")
        email_soporte = request.POST.get("email_soporte")

        # Aquí podrías guardar los valores en un modelo Configuracion o archivo JSON
        message = f"✅ Configuración actualizada: {nombre_sistema}, moneda {moneda}, soporte {email_soporte}"

    context = {
        "message": message
    }
    return render(request, "finanzas/admin_config.html", context)


# --- ADMIN LOGS ---
@login_required
@user_passes_test(lambda u: u.is_staff)
def admin_logs(request):
    movimientos = MovimientoLog.objects.select_related("usuario").all().order_by("-fecha")
    logs = SystemLog.objects.select_related("usuario").all()

    # Filtro opcional por usuario
    usuario_filtro = request.GET.get("usuario")
    if usuario_filtro:
        logs = logs.filter(usuario__username__icontains=usuario_filtro)

    return render(request, "finanzas/admin_logs.html", {
        "logs": logs, 
        "movimientos": movimientos,
        "usuario_filtro": usuario_filtro
        })

def registrar_log(usuario, accion, detalle=""):
    """Guarda una acción administrativa o de usuario en la tabla de logs."""
    SystemLog.objects.create(usuario=usuario, accion=accion, detalle=detalle)